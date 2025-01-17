import random
import openpyxl
import os

class PonyCharacterPromptPicker:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": os.path.join("pony_char_list.xlsx")}),
                "sheet_name": ("STRING", {"default": "Female character list"}),
                "column_letter": ("STRING", {"default": "A"}),
                "prefix": ("STRING", {"default": "score_9, score_8_up, score_7_up"}),
                "seed_mode": (["randomize", "fixed"],),
                "seed": ("INT", {"default": 0, "min": 0, "max": 0xffffffffffffffff}),
            },
        }

    RETURN_TYPES = ("STRING", "INT")
    RETURN_NAMES = ("prompt", "seed")
    FUNCTION = "pick_prompt"
    CATEGORY = "prompt"

    def pick_prompt(self, excel_path, sheet_name, column_letter, prefix, seed_mode, seed):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_excel_path = os.path.join(current_dir, excel_path)
        
        if not os.path.exists(full_excel_path):
            raise FileNotFoundError(f"Excel file not found: {full_excel_path}")

        workbook = openpyxl.load_workbook(full_excel_path, read_only=True)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")

        sheet = workbook[sheet_name]

        # Get the column index from the letter
        column_index = openpyxl.utils.column_index_from_string(column_letter)

        # Collect valid cell values
        valid_cells = []
        for row in sheet.iter_rows(min_row=3, min_col=column_index, max_col=column_index):
            cell = row[0]  # There's only one cell per row since we specified min_col=max_col
            if cell.value is not None:
                valid_cells.append(cell.value)

        workbook.close()

        if not valid_cells:
            raise ValueError(f"No valid values found in column {column_letter} starting from row 3")

        # Set the seed for random selection
        if seed_mode == "randomize":
            seed = random.randint(0, 0xffffffffffffffff)
        random.seed(seed)

        selected_prompt = random.choice(valid_cells)

        # Process the prefix
        prefix_list = [p.strip() for p in prefix.split(',')]
        
        # Add comma and space to the selected prompt if needed
        if not selected_prompt.endswith(", "):
            selected_prompt += ", "

        # Combine prefix and selected prompt
        full_prompt = ", ".join(prefix_list) + ", " + selected_prompt

        return (full_prompt, seed)